from django.views.generic import ListView, DetailView, UpdateView, CreateView
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.urls import reverse, reverse_lazy
from django.http import HttpResponse
from django.contrib import messages
from django.views import View
from django.db.models import Q
from .models import *
from .forms import * 

from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
 
class Login(View):

    def get(self, request):
        return render(request, 'login.html')

    def post(self, request):
        data = request.POST
        username = data.get('username')
        password = data.get('password')

        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            return redirect('home')
        else:
            return render(request, 'login.html', {})

class Logout(View):
    def get(self, request):
        logout(request)
        return redirect('login')

@method_decorator(login_required, name='dispatch')
class Home(View):

    def get(self, request):
        print(request.user)

        if request.user.systemuser.role == "attorney":
            user_role = 'attorney'
            system_user = Attorney.objects.get(user=request.user.systemuser)

        elif request.user.systemuser.role == "assistant":
            user_role = 'assistant'
            system_user = Assistant.objects.get(user=request.user.systemuser)

        else:
            user_role = 'admin'
            system_user = request.user.systemuser

        return render(request, 'home.html', {'sysuser':system_user, 'role':user_role}) 

def CaseContext(request, form_class=CaseForm(), return_query=True):
    search = request.GET.get('search', '')
    assistant_id = request.GET.get('assistant', '')
    attorney_id = request.GET.get('attorney', '')
    status_id = request.GET.get('status', '')
    claim_status = request.GET.get('claim_status', '')
    
    context = {
        'name': 'Case Registers',
        'form': form_class,
        'search': search,
        'delete_active': False,
        'restore_active': False,
        # Datos para los selects de filtro
        'assistants': Assistant.objects.filter(active=True).order_by('name'),
        'attorneys': Attorney.objects.filter(active=True).order_by('name'),
        'statuses': CaseStatus.objects.filter(active=True).order_by('name'),
        'statuses': CaseStatus.objects.filter(active=True),
        # Valores seleccionados actualmente
        'selected_assistant': assistant_id,
        'selected_attorney': attorney_id,
        'selected_status': status_id,
        'selected_claim_status': claim_status,
    }
    
    if not return_query:
        return context
    
    # Base queryset
    queryset = Case.objects.filter(
        Q(client__name__icontains=search) | Q(employer__name__icontains=search)
    )
    
    # Aplicar filtros
    if assistant_id:
        queryset = queryset.filter(assistant_id=assistant_id)
    
    if attorney_id:
        queryset = queryset.filter(attorney_id=attorney_id)
    
    if status_id:
        queryset = queryset.filter(status_id=status_id)
    
    if claim_status:
        queryset = queryset.filter(claim_status=claim_status)
    
    # Filtros por rol del usuario
    if request.user.systemuser.role != "admin":
        queryset = queryset.filter(active=True)
    
    if request.user.systemuser.role == "attorney":
        context['user_is_attorney'] = True
        attorney = Attorney.objects.get(user=request.user.systemuser)
        queryset = queryset.filter(attorney=attorney)
    
    if request.user.systemuser.role == "assistant":
        context['user_is_assistant'] = True
        assistant = Assistant.objects.get(user=request.user.systemuser)
        queryset = queryset.filter(assistant=assistant)
    
    context['items'] = queryset
    
    return context


@method_decorator(login_required, name='dispatch')
class CreateCase(View):
    
    def get(self, request):
        # GET request: formularios vacíos
        case_form = CaseForm(prefix='case')
        client_form = ClientChoiceForm(prefix='client')
        employer_form = EmployerChoiceForm(prefix='employer')

        # Usar los forms opcionales SIN instancia (None explícito)
        insurance_form = InsuranceCarrierChoiceForm(prefix='insurance', instance=None)
        def_lawfirm_form = DefenseLawFirmChoiceForm(prefix='def_lawfirm', instance=None)
        def_attorney_form = DefenseAttorneyChoiceForm(prefix='def_attorney', instance=None)
        def_assistant_form = DefenseAssistantChoiceForm(prefix='def_assistant', instance=None)
        claim_admin_form = ClaimAdministratorChoiceForm(prefix='claim_admin', instance=None)
        claim_adjuster_form = ClaimAdjusterChoiceForm(prefix='claim_adjuster', instance=None)
 
        client_contact_formset = ClientContactFormSet(prefix='client_contacts')
        injury_formset = InjuryFormSet(prefix='injuries')
        claim_adjuster_contact_formset = ClaimAdjusterContactFormSet(prefix='claim_adjuster_contacts')
        defense_attorney_contact_formset = DefenseAttorneyContactFormSet(prefix='def_attorney_contacts')
        defense_assistant_contact_formset = DefenseAssistantContactFormSet(prefix='def_assistant_contacts')
    
        context = {
            'case_form': case_form,
            'client_form': client_form,
            'employer_form': employer_form,
            'insurance_form': insurance_form,
            'def_lawfirm_form': def_lawfirm_form,
            'def_attorney_form': def_attorney_form,
            'def_assistant_form': def_assistant_form,
            'claim_admin_form': claim_admin_form,
            'claim_adjuster_form': claim_adjuster_form,
            'injury_formset': injury_formset,
            'client_contact_formset': client_contact_formset,
            'claim_adjuster_contact_formset': claim_adjuster_contact_formset,
            'def_attorney_contact_formset': defense_attorney_contact_formset,
            'def_assistant_contact_formset': defense_assistant_contact_formset,
            'create_form':1,
        }
        return render(request, 'create-case.html', context)

    def post(self, request):
        data = request.POST
        
        # Formulario principal del Case
        case_form = CaseForm(data, prefix='case')
        
        # Formsets anidados
        client_contact_formset = ClientContactFormSet(data, prefix='client_contacts')
        injury_formset = InjuryFormSet(data, prefix='injuries')
        
        # Forms OBLIGATORIOS
        client_form = ClientChoiceForm(data, prefix='client')
        employer_form = EmployerChoiceForm(data, prefix='employer')
        
        # ✅ CORREGIDO: Forms OPCIONALES con instance=None explícito
        insurance_form = InsuranceCarrierChoiceForm(data, prefix='insurance', instance=None)
        def_lawfirm_form = DefenseLawFirmChoiceForm(data, prefix='def_lawfirm', instance=None)
        def_attorney_form = DefenseAttorneyChoiceForm(data, prefix='def_attorney', instance=None)
        def_assistant_form = DefenseAssistantChoiceForm(data, prefix='def_assistant', instance=None)
        claim_admin_form = ClaimAdministratorChoiceForm(data, prefix='claim_admin', instance=None)
        claim_adjuster_form = ClaimAdjusterChoiceForm(data, prefix='claim_adjuster', instance=None)
        
        # Formsets para contactos
        claim_adjuster_contact_formset = ClaimAdjusterContactFormSet(
            data, prefix='claim_adjuster_contacts'
        )
        defense_attorney_contact_formset = DefenseAttorneyContactFormSet(
            data, prefix='def_attorney_contacts'
        )
        defense_assistant_contact_formset = DefenseAssistantContactFormSet(
            data, prefix='def_assistant_contacts'
        )
        
        # Validar TODOS los forms (obligatorios + opcionales + formsets)
        # Los opcionales se validan aunque estén vacíos
        all_forms_valid = all([
            case_form.is_valid(),
            client_form.is_valid(),
            employer_form.is_valid(),
            insurance_form.is_valid(),
            def_lawfirm_form.is_valid(),
            def_attorney_form.is_valid(),
            def_assistant_form.is_valid(),
            claim_admin_form.is_valid(),
            claim_adjuster_form.is_valid(),
            injury_formset.is_valid(),
            client_contact_formset.is_valid(),
            claim_adjuster_contact_formset.is_valid(),
            defense_attorney_contact_formset.is_valid(),
            defense_assistant_contact_formset.is_valid(),
        ])

        
        if all_forms_valid:
            # Guardar instancias OBLIGATORIAS
            client = client_form.save() 

            print(client) 
            employer = employer_form.save()

            print(employer)
            
            # Guardar Case con las relaciones obligatorias
            case = case_form.save(commit=False)
            case.client = client
            case.employer = employer
            
            # ✅ CORREGIDO: Procesar forms OPCIONALES
            # El OptionalModelForm.save() debe retornar None si no hay datos
            insurance = insurance_form.save() if insurance_form.has_changed() else None
            def_lawfirm = def_lawfirm_form.save() if def_lawfirm_form.has_changed() else None
            def_attorney = def_attorney_form.save() if def_attorney_form.has_changed() else None
            def_assistant = def_assistant_form.save() if def_assistant_form.has_changed() else None
            claim_admin = claim_admin_form.save() if claim_admin_form.has_changed() else None
            claim_adjuster = claim_adjuster_form.save() if claim_adjuster_form.has_changed() else None
            
            # Asignar solo si no son None
            if insurance:
                case.insurance = insurance
            if def_lawfirm:
                case.def_lawfirm = def_lawfirm
            if def_attorney:
                case.def_attorney = def_attorney
            if def_assistant:
                case.def_assistant = def_assistant
            if claim_admin:
                case.claim_admin = claim_admin
            if claim_adjuster:
                case.claim_adjuster = claim_adjuster
            
            case.save()
            
            # Guardar formsets que dependen de Client (siempre)
            client_contact_formset.instance = client
            client_contact_formset.save()
            
            # Guardar injury_formset
            injury_formset.instance = case
            injury_formset.save()
            
            # ✅ CORREGIDO: Guardar formsets solo si las instancias existen
            if claim_adjuster:
                claim_adjuster_contact_formset.instance = claim_adjuster
                claim_adjuster_contact_formset.save()
            
            if def_attorney:
                defense_attorney_contact_formset.instance = def_attorney
                defense_attorney_contact_formset.save()
            
            if def_assistant:
                defense_assistant_contact_formset.instance = def_assistant
                defense_assistant_contact_formset.save()
            
            return redirect('case')
        
        # Si hay errores, devolver el template con los datos ingresados
        context = {
            'case_form': case_form,
            'client_form': client_form,
            'employer_form': employer_form,
            'insurance_form': insurance_form,
            'def_lawfirm_form': def_lawfirm_form,
            'def_attorney_form': def_attorney_form,
            'def_assistant_form': def_assistant_form,
            'claim_admin_form': claim_admin_form,
            'claim_adjuster_form': claim_adjuster_form,
            'injury_formset': injury_formset,
            'client_contact_formset': client_contact_formset,
            'claim_adjuster_contact_formset': claim_adjuster_contact_formset,
            'def_attorney_contact_formset': defense_attorney_contact_formset,
            'def_assistant_contact_formset': defense_assistant_contact_formset,
            'create_form':1,
        }
        return render(request, 'create-case.html', context)
 
@method_decorator(login_required, name='dispatch')
class UpdateCase(View):
    
    def get(self, request, pk):
        case = get_object_or_404(Case, id=pk)
        
        client = case.client
        employer = case.employer
        insurance = case.insurance
        def_lawfirm = case.def_lawfirm
        def_attorney = case.def_attorney
        def_assistant = case.def_assistant
        claim_admin = case.claim_admin
        claim_adjuster = case.claim_adjuster
        
        # Forms normales (siempre con instance)
        case_form = CaseForm(prefix='case', instance=case)
        client_form = ClientForm(prefix='client', instance=client)
        employer_form = EmployerForm(prefix='employer', instance=employer)
        
        # Forms opcionales: normal si existe, choice si no
        insurance_form = InsuranceCarrierForm(prefix='insurance', instance=insurance) if insurance else InsuranceCarrierChoiceForm(prefix='insurance')
        def_lawfirm_form = DefenseLawFirmForm(prefix='def_lawfirm', instance=def_lawfirm) if def_lawfirm else DefenseLawFirmChoiceForm(prefix='def_lawfirm')
        def_attorney_form = DefenseAttorneyForm(prefix='def_attorney', instance=def_attorney) if def_attorney else DefenseAttorneyChoiceForm(prefix='def_attorney')
        def_assistant_form = DefenseAssistantForm(prefix='def_assistant', instance=def_assistant) if def_assistant else DefenseAssistantChoiceForm(prefix='def_assistant')
        claim_admin_form = ClaimAdministratorForm(prefix='claim_admin', instance=claim_admin) if claim_admin else ClaimAdministratorChoiceForm(prefix='claim_admin')
        claim_adjuster_form = ClaimAdjusterForm(prefix='claim_adjuster', instance=claim_adjuster) if claim_adjuster else ClaimAdjusterChoiceForm(prefix='claim_adjuster')
        
        # Formsets
        injury_formset = InjuryFormSet(prefix='injuries', instance=case, queryset=case.injuries.filter(active=True))
        client_contact_formset = ClientContactFormSet(prefix='client_contacts', instance=client, queryset=client.contacts.filter(active=True))
        claim_adjuster_contact_formset = ClaimAdjusterContactFormSet(prefix='claim_adjuster_contacts', instance=claim_adjuster, queryset=claim_adjuster.contacts.filter(active=True)) if claim_adjuster else ClaimAdjusterContactFormSet(prefix='claim_adjuster_contacts')
        defense_attorney_contact_formset = DefenseAttorneyContactFormSet(prefix='def_attorney_contacts', instance=def_attorney, queryset=def_attorney.contacts.filter(active=True)) if def_attorney else DefenseAttorneyContactFormSet(prefix='def_attorney_contacts')
        defense_assistant_contact_formset = DefenseAssistantContactFormSet(prefix='def_assistant_contacts', instance=def_assistant, queryset=def_assistant.contacts.filter(active=True)) if def_assistant else DefenseAssistantContactFormSet(prefix='def_assistant_contacts')
        
        context = {
            'case_form': case_form,
            'client_form': client_form,
            'employer_form': employer_form,
            'insurance_form': insurance_form,
            'def_lawfirm_form': def_lawfirm_form,
            'def_attorney_form': def_attorney_form,
            'def_assistant_form': def_assistant_form,
            'claim_admin_form': claim_admin_form,
            'claim_adjuster_form': claim_adjuster_form,
            'injury_formset': injury_formset,
            'client_contact_formset': client_contact_formset,
            'claim_adjuster_contact_formset': claim_adjuster_contact_formset,
            'def_attorney_contact_formset': defense_attorney_contact_formset,
            'def_assistant_contact_formset': defense_assistant_contact_formset,
        }
        return render(request, 'create-case.html', context)
    
    def post(self, request, pk):
        data = request.POST
        case = get_object_or_404(Case, id=pk)
        
        # Obtener instancias existentes
        client = case.client
        employer = case.employer
        insurance = case.insurance
        def_lawfirm = case.def_lawfirm
        def_attorney = case.def_attorney
        def_assistant = case.def_assistant
        claim_admin = case.claim_admin
        claim_adjuster = case.claim_adjuster
        
        # Formulario principal
        case_form = CaseForm(data, prefix='case', instance=case)
        
        # Forms OBLIGATORIOS
        client_form = ClientForm(data, prefix='client', instance=client)
        employer_form = EmployerForm(data, prefix='employer', instance=employer)
        
        # ✅ MISMA LÓGICA QUE GET: normal si existe, choice si no
        insurance_form = InsuranceCarrierForm(data, prefix='insurance', instance=insurance) if insurance else InsuranceCarrierChoiceForm(data, prefix='insurance')
        def_lawfirm_form = DefenseLawFirmForm(data, prefix='def_lawfirm', instance=def_lawfirm) if def_lawfirm else DefenseLawFirmChoiceForm(data, prefix='def_lawfirm')
        def_attorney_form = DefenseAttorneyForm(data, prefix='def_attorney', instance=def_attorney) if def_attorney else DefenseAttorneyChoiceForm(data, prefix='def_attorney')
        def_assistant_form = DefenseAssistantForm(data, prefix='def_assistant', instance=def_assistant) if def_assistant else DefenseAssistantChoiceForm(data, prefix='def_assistant')
        claim_admin_form = ClaimAdministratorForm(data, prefix='claim_admin', instance=claim_admin) if claim_admin else ClaimAdministratorChoiceForm(data, prefix='claim_admin')
        claim_adjuster_form = ClaimAdjusterForm(data, prefix='claim_adjuster', instance=claim_adjuster) if claim_adjuster else ClaimAdjusterChoiceForm(data, prefix='claim_adjuster')
        
        # Formsets
        injury_formset = InjuryFormSet(data, prefix='injuries', instance=case, queryset=case.injuries.all())
        client_contact_formset = ClientContactFormSet(data, prefix='client_contacts', instance=client, queryset=client.contacts.all())
        
        # Formsets opcionales
        claim_adjuster_contact_formset = None
        defense_attorney_contact_formset = None
        defense_assistant_contact_formset = None
        
        if claim_adjuster:
            claim_adjuster_contact_formset = ClaimAdjusterContactFormSet(
                data, prefix='claim_adjuster_contacts', 
                instance=claim_adjuster, 
                queryset=claim_adjuster.contacts.all()
            )
        else:
            claim_adjuster_contact_formset = ClaimAdjusterContactFormSet(data, prefix='claim_adjuster_contacts')
        
        if def_attorney:
            defense_attorney_contact_formset = DefenseAttorneyContactFormSet(
                data, prefix='def_attorney_contacts', 
                instance=def_attorney, 
                queryset=def_attorney.contacts.all()
            )
        else:
            defense_attorney_contact_formset = DefenseAttorneyContactFormSet(data, prefix='def_attorney_contacts')
        
        if def_assistant:
            defense_assistant_contact_formset = DefenseAssistantContactFormSet(
                data, prefix='def_assistant_contacts', 
                instance=def_assistant, 
                queryset=def_assistant.contacts.all()
            )
        else:
            defense_assistant_contact_formset = DefenseAssistantContactFormSet(data, prefix='def_assistant_contacts')
        
        # ✅ VALIDAR TODOS LOS FORMS
        all_forms_valid = all([
            case_form.is_valid(),
            client_form.is_valid(),
            employer_form.is_valid(),
            insurance_form.is_valid(),
            def_lawfirm_form.is_valid(),
            def_attorney_form.is_valid(),
            def_assistant_form.is_valid(),
            claim_admin_form.is_valid(),
            claim_adjuster_form.is_valid(),
            injury_formset.is_valid(),
            client_contact_formset.is_valid(),
            claim_adjuster_contact_formset.is_valid(),
            defense_attorney_contact_formset.is_valid(),
            defense_assistant_contact_formset.is_valid(),
        ])
        
        if all_forms_valid:
            # Guardar obligatorios
            client = client_form.save()
            employer = employer_form.save()
            
            # ✅ FUNCIÓN HELPER para guardar forms opcionales
            def save_optional_form(form, existing_instance):
                """Guarda un form opcional (puede ser ModelForm o ChoiceForm)"""
                if form.has_changed():
                    # Si es ChoiceForm, tiene cleaned_data con _existing_object
                    # Si es ModelForm normal, guarda normalmente
                    if hasattr(form, 'cleaned_data') and '_existing_object' in form.cleaned_data:
                        return form.cleaned_data['_existing_object']
                    return form.save()
                return existing_instance
            
            # Guardar opcionales
            new_insurance = save_optional_form(insurance_form, insurance)
            new_def_lawfirm = save_optional_form(def_lawfirm_form, def_lawfirm)
            new_def_attorney = save_optional_form(def_attorney_form, def_attorney)
            new_def_assistant = save_optional_form(def_assistant_form, def_assistant)
            new_claim_admin = save_optional_form(claim_admin_form, claim_admin)
            new_claim_adjuster = save_optional_form(claim_adjuster_form, claim_adjuster)
            
            # Actualizar Case
            case = case_form.save(commit=False)
            case.client = client
            case.employer = employer
            case.insurance = new_insurance
            case.def_lawfirm = new_def_lawfirm
            case.def_attorney = new_def_attorney
            case.def_assistant = new_def_assistant
            case.claim_admin = new_claim_admin
            case.claim_adjuster = new_claim_adjuster
            case.save()
            
            # Guardar formsets
            injury_formset.save()
            client_contact_formset.save()
            
            # Guardar formsets opcionales
            if new_claim_adjuster:
                claim_adjuster_contact_formset.instance = new_claim_adjuster
                claim_adjuster_contact_formset.save()
            
            if new_def_attorney:
                defense_attorney_contact_formset.instance = new_def_attorney
                defense_attorney_contact_formset.save()
            
            if new_def_assistant:
                defense_assistant_contact_formset.instance = new_def_assistant
                defense_assistant_contact_formset.save()
            
            return redirect('case')
        
        # Si hay errores, devolver template
        context = {
            'case_form': case_form,
            'client_form': client_form,
            'employer_form': employer_form,
            'insurance_form': insurance_form,
            'def_lawfirm_form': def_lawfirm_form,
            'def_attorney_form': def_attorney_form,
            'def_assistant_form': def_assistant_form,
            'claim_admin_form': claim_admin_form,
            'claim_adjuster_form': claim_adjuster_form,
            'injury_formset': injury_formset,
            'client_contact_formset': client_contact_formset,
            'claim_adjuster_contact_formset': claim_adjuster_contact_formset,
            'def_attorney_contact_formset': defense_attorney_contact_formset,
            'def_assistant_contact_formset': defense_assistant_contact_formset,
        }
        return render(request, 'create-case.html', context)
@method_decorator(login_required, name='dispatch')
class DetailCase(DetailView):
    model = Case
    template_name = 'case-detail.html'
    context_object_name = 'case'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_injuries'] = self.object.injuries.filter(active=True)
        return context  

@method_decorator(login_required, name='dispatch')
class DeleteCase(View):
    def post(self, request, pk):

        print("que hace?")
        case = Case.objects.get(id=pk)

        case.soft_delete()
        case.save()

        for injury in case.injuries.all():
            injury.soft_delete()
            injury.save()

        return redirect('case')

@method_decorator(login_required, name='dispatch')
class RestoreCase(View):
    def post(self, request, pk):
        case = Case.objects.get(id=pk)

        case.restore()
        case.save()

        for injury in case.injuries.all():
            injury.restore()
            injury.save()

        return redirect('case')

@method_decorator(login_required, name='dispatch')
class ListCases(ListView):
    model = Case
    template_name = 'list-cases.html'
    context_object_name = 'items'
    paginate_by = 20  # Opcional: paginación

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['detail_url'] = True
        context['list_url'] = 'case'
        
        # Agregar contexto con filtros
        filter_context = CaseContext(self.request)
        context.update(filter_context)
        
        return context

 
# ============================================================
# BASES REUTILIZABLES (definidas aquí mismo para el ejemplo)
# ============================================================

@method_decorator(login_required, name='dispatch')
class BaseListView(ListView):
    # ListView base con contexto automático
    template_name = None
    context_object_name = 'items'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if hasattr(self, 'get_context_function'):
            added_context = self.get_context_function(self.request)
            context.update(added_context)
        return context


@method_decorator(login_required, name='dispatch')
class BaseCreateView(CreateView):
    # CreateView base con manejo automático 
    template_name = None
    success_message = "registro creado exitosamente"

    def get(self, request, *args, **kwargs):
        # Maneja la petición GET - muestra el formulario vacío 
        form = self.get_form()
        context = self.get_context_function(request, form)
        context['create_active'] = True
        return self.render_to_response(context)
    
    def post(self, request, *args, **kwargs):
        # Maneja la petición POST - procesa el formulario 
        form = self.get_form()
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        context = self.get_context_function(self.request, form)
        context['create_active'] = True
        return self.render_to_response(context)
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, self.success_message)
        return response


@method_decorator(login_required, name='dispatch')
class BaseUpdateView(UpdateView):
    # UpdateView base con manejo automático
    template_name = None
    success_message = "registro modificado exitosamente"

    def get(self, request, *args, **kwargs):
        # Maneja la petición GET - muestra el formulario con datos del objeto
        self.object = self.get_object()
        form = self.get_form() 
        context = self.get_context_function(request, form)
        context['update_active'] = True
        context['temp_name'] = self.object
        return self.render_to_response(context)
    
    def post(self, request, *args, **kwargs):
        # Maneja la petición POST - procesa el formulario
        self.object = self.get_object()
        form = self.get_form()

        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        context = self.get_context_function(self.request, form)
        context['update_active'] = True
        context['temp_name'] = self.get_object()
        return self.render_to_response(context)
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, self.success_message)
        return response 

@method_decorator(login_required, name='dispatch')
class BaseDeleteView(View):

    # DeleteView base
    model = None
    redirect_url = None  # URL de lista, ej: 'status'
    success_message = "registro eliminado exitosamente"
    
    def has_related_cases(self, obj):
        """Verifica si el objeto tiene casos relacionados"""
        if hasattr(obj, 'cases'):
            if obj.cases.filter(active=True).exists():
                return True
        return False
    
    def get_related_cases_count(self, obj):
        """Retorna el número de casos relacionados"""
        if hasattr(obj, 'cases'):
            return obj.cases.filter(active=True).count()
        return 0
    
    def get(self, request, pk):
        """Maneja GET - muestra la lista con el modal activo"""
        obj = self.model.objects.get(id=pk)

        context = self.get_context_function(request)
        context['delete_active'] = True
        context['temp_name'] = obj.name 

        # Verificar si tiene casos relacionados
        if self.has_related_cases(obj):
            count = self.get_related_cases_count(obj)
            error_msg = f"Cannot delete '{obj.name}'. It has {count} case(s) related"
            context['error_msg'] = error_msg
            return render(request, self.template_name, context )
         
        return redirect(self.redirect_url)
    
    def post(self, request, pk):
        obj = self.model.objects.get(id=pk)
        
        # Verificar nuevamente por si acaso
        if self.has_related_cases(obj):
            count = self.get_related_cases_count(obj)
            error_msg = f"Cannot delete '{obj.name}'. It has {count} case(s) related"
            messages.error(request, error_msg)
            # ✅ Redirigir al GET (misma URL) para mostrar el modal con error
            return redirect(request.path)
        
        # Proceder con el soft delete
        obj.soft_delete()
        messages.success(request, self.success_message)
        return redirect(self.redirect_url)


@method_decorator(login_required, name='dispatch')
class BaseRestoreView(View):
    # RestoreView base
    model = None
    redirect_url = None
    success_message = "registro restaurado exitosamente"
    
    def post(self, request, pk):
        obj = self.model.objects.get(id=pk)
        obj.restore()
        messages.success(request, self.success_message)
        return redirect(self.redirect_url)


# ============================================================
# CASE STATUS
# ============================================================

def case_status_context(request, form=CaseStatusForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Case Status',
        'form': form,
        'list_url': 'status',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = CaseStatus.objects.filter(
            Q(name__icontains=search) | Q(id__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListCaseStatus(BaseListView):
    model = CaseStatus
    template_name = 'show/show-status.html'
    
    def get_context_function(self, request):
        return case_status_context(request)


@method_decorator(login_required, name='dispatch')
class CreateCaseStatus(BaseCreateView):
    model = CaseStatus
    form_class = CaseStatusForm
    success_url = reverse_lazy('status')
    template_name = 'show/show-status.html'
    success_message = "Case Status creado exitosamente"
    
    def get_context_function(self, request, form):
        
        return case_status_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateCaseStatus(BaseUpdateView):
    model = CaseStatus
    form_class = CaseStatusForm
    success_url = reverse_lazy('status')
    template_name = 'show/show-status.html'
    success_message = "Case Status modificado exitosamente"
    
    def get_context_function(self, request, form):
        return case_status_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteCaseStatus(BaseDeleteView):
    model = CaseStatus
    redirect_url = 'status'
    template_name = 'show/show-status.html'
    success_message = "Case Status eliminado exitosamente"

    def get_context_function(self, request):
        return case_status_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreCaseStatus(BaseRestoreView):
    model = CaseStatus
    redirect_url = 'status'
    success_message = "Case Status restaurado exitosamente"


# ============================================================
# ATTORNEY
# ============================================================

def attorney_context(request, form=SystemAttorneyForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Attorney',
        'form': form,
        'list_url': 'attorney',
        'search': search,
        'form': SystemAttorneyForm,
        'update_form': SystemAttorneyUpdateForm,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = Attorney.objects.filter(
            Q(name__icontains=search) | Q(id__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListAttorney(BaseListView):
    model = Attorney
    template_name = 'show/show-attorney.html'
    
    def get_context_function(self, request):
        return attorney_context(request)


@method_decorator(login_required, name='dispatch')
class CreateAttorney(BaseCreateView):
    model = Attorney
    form_class = SystemAttorneyForm
    success_url = reverse_lazy('attorney')
    template_name = 'show/show-attorney.html'
    success_message = "Attorney creado exitosamente"
    
    def get_context_function(self, request, form):
        return attorney_context(request, form) 


@method_decorator(login_required, name='dispatch')
class UpdateAttorney(BaseUpdateView):
    model = Attorney
    form_class = SystemAttorneyUpdateForm
    success_url = reverse_lazy('attorney')
    template_name = 'show/show-attorney.html'
    success_message = "Attorney modificado exitosamente"
    
    def get_context_function(self, request, form):
        return attorney_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteAttorney(BaseDeleteView):
    model = Attorney
    redirect_url = 'attorney'
    template_name = 'show/show-attorney.html'
    success_message = "Attorney eliminado exitosamente"

    def get_context_function(self, request):
        return attorney_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreAttorney(BaseRestoreView):
    model = Attorney
    redirect_url = 'attorney'
    success_message = "Attorney restaurado exitosamente"


# ============================================================
# ASSISTANT
# ============================================================

def assistant_context(request, form=SystemAssistantForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Assistant',
        'form': form,
        'list_url': 'assistant',
        'search': search,
        'update_form': SystemAssistantUpdateForm,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = Assistant.objects.filter(
            Q(name__icontains=search) | Q(id__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListAssistant(BaseListView):
    model = Assistant
    template_name = 'show/show-assistant.html'
    
    def get_context_function(self, request):
        return assistant_context(request)


@method_decorator(login_required, name='dispatch')
class CreateAssistant(BaseCreateView):
    model = Assistant
    form_class = SystemAssistantForm
    success_url = reverse_lazy('assistant')
    template_name = 'show/show-assistant.html'
    success_message = "Assistant creado exitosamente"
    
    def get_context_function(self, request, form):
        return assistant_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateAssistant(BaseUpdateView):
    model = Assistant
    form_class = SystemAssistantUpdateForm
    success_url = reverse_lazy('assistant')
    template_name = 'show/show-assistant.html'
    success_message = "Assistant modificado exitosamente"
    
    def get_context_function(self, request, form):
        return assistant_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteAssistant(BaseDeleteView):
    model = Assistant
    redirect_url = 'assistant'
    template_name = 'show/show-assistant.html'
    success_message = "Assistant eliminado exitosamente"

    def get_context_function(self, request):
        return assistant_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreAssistant(BaseRestoreView):
    model = Assistant
    redirect_url = 'assistant'
    success_message = "Assistant restaurado exitosamente"


# ============================================================
# EMPLOYER
# ============================================================

def employer_context(request, form=EmployerForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Employer',
        'form': form,
        'list_url': 'employer',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = Employer.objects.filter(
            Q(name__icontains=search) | 
            Q(id__icontains=search) |
            Q(address__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListEmployer(BaseListView):
    model = Employer
    template_name = 'show/show-employer.html'
    
    def get_context_function(self, request):
        return employer_context(request)


@method_decorator(login_required, name='dispatch')
class CreateEmployer(BaseCreateView):
    model = Employer
    form_class = EmployerForm
    success_url = reverse_lazy('employer')
    template_name = 'show/show-employer.html'
    success_message = "Employer creado exitosamente"
    
    def get_context_function(self, request, form):
        return employer_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateEmployer(BaseUpdateView):
    model = Employer
    form_class = EmployerForm
    success_url = reverse_lazy('employer')
    template_name = 'show/show-employer.html'
    success_message = "Employer modificado exitosamente"
    
    def get_context_function(self, request, form):
        return employer_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteEmployer(BaseDeleteView):
    model = Employer
    redirect_url = 'employer'
    template_name = 'show/show-employer.html'
    success_message = "Employer eliminado exitosamente"

    def get_context_function(self, request):
        return employer_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreEmployer(BaseRestoreView):
    model = Employer
    redirect_url = 'employer'
    success_message = "Employer restaurado exitosamente"


# ============================================================
# INSURANCE CARRIER
# ============================================================

def insurance_carrier_context(request, form=InsuranceCarrierForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Insurance Carrier',
        'form': form,
        'list_url': 'insurance-carrier',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = InsuranceCarrier.objects.filter(
            Q(name__icontains=search) | 
            Q(id__icontains=search) |
            Q(address__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListInsuranceCarrier(BaseListView):
    model = InsuranceCarrier
    template_name = 'show/show-insurance-carrier.html'
    
    def get_context_function(self, request):
        return insurance_carrier_context(request)


@method_decorator(login_required, name='dispatch')
class CreateInsuranceCarrier(BaseCreateView):
    model = InsuranceCarrier
    form_class = InsuranceCarrierForm
    success_url = reverse_lazy('insurance-carrier')
    template_name = 'show/show-insurance-carrier.html'
    success_message = "Insurance Carrier creado exitosamente"
    
    def get_context_function(self, request, form):
        return insurance_carrier_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateInsuranceCarrier(BaseUpdateView):
    model = InsuranceCarrier
    form_class = InsuranceCarrierForm
    success_url = reverse_lazy('insurance-carrier')
    template_name = 'show/show-insurance-carrier.html'
    success_message = "Insurance Carrier modificado exitosamente"
    
    def get_context_function(self, request, form):
        return insurance_carrier_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteInsuranceCarrier(BaseDeleteView):
    model = InsuranceCarrier
    redirect_url = 'insurance-carrier'
    template_name = 'show/show-insurance-carrier.html'
    success_message = "Insurance Carrier eliminado exitosamente"

    def get_context_function(self, request):
        return insurance_carrier_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreInsuranceCarrier(BaseRestoreView):
    model = InsuranceCarrier
    redirect_url = 'insurance-carrier'
    success_message = "Insurance Carrier restaurado exitosamente"


# ============================================================
# CLAIM ADMINISTRATOR
# ============================================================

def claim_administrator_context(request, form=ClaimAdministratorForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Claim Administrator',
        'form': form,
        'list_url': 'claim-administrator',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = ClaimAdministrator.objects.filter(
            Q(name__icontains=search) | 
            Q(id__icontains=search) |
            Q(address__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    return context


@method_decorator(login_required, name='dispatch')
class ListClaimAdministrator(BaseListView):
    model = ClaimAdministrator
    template_name = 'show/show-claim-administrator.html'
    
    def get_context_function(self, request):
        return claim_administrator_context(request)


@method_decorator(login_required, name='dispatch')
class CreateClaimAdministrator(BaseCreateView):
    model = ClaimAdministrator
    form_class = ClaimAdministratorForm
    success_url = reverse_lazy('claim-administrator')
    template_name = 'show/show-claim-administrator.html'
    success_message = "Claim Administrator creado exitosamente"
    
    def get_context_function(self, request, form):
        return claim_administrator_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateClaimAdministrator(BaseUpdateView):
    model = ClaimAdministrator
    form_class = ClaimAdministratorForm
    success_url = reverse_lazy('claim-administrator')
    template_name = 'show/show-claim-administrator.html'
    success_message = "Claim Administrator modificado exitosamente"
    
    def get_context_function(self, request, form):
        return claim_administrator_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteClaimAdministrator(BaseDeleteView):
    model = ClaimAdministrator
    redirect_url = 'claim-administrator'
    template_name = 'show/show-claim_administrator.html'
    success_message = "Claim Administrator eliminado exitosamente"


@method_decorator(login_required, name='dispatch')
class RestoreClaimAdministrator(BaseRestoreView):
    model = ClaimAdministrator
    redirect_url = 'claim-administrator'
    success_message = "Claim Administrator restaurado exitosamente"

    def get_context_function(self, request):
        return claim_administrator_context(request)


# ============================================================
# DEFENSE LAW FIRM
# ============================================================

def defense_law_firm_context(request, form=DefenseLawFirmForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Defense Law Firm',
        'form': form,
        'list_url': 'defense-law-firm',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = DefenseLawFirm.objects.filter(
            Q(name__icontains=search) | 
            Q(id__icontains=search) |
            Q(address__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListDefenseLawFirm(BaseListView):
    model = DefenseLawFirm
    template_name = 'show/show-defense-law-firm.html'
    
    def get_context_function(self, request):
        return defense_law_firm_context(request)


@method_decorator(login_required, name='dispatch')
class CreateDefenseLawFirm(BaseCreateView):
    model = DefenseLawFirm
    form_class = DefenseLawFirmForm
    success_url = reverse_lazy('defense-law-firm')
    template_name = 'show/show-defense-law-firm.html'
    success_message = "Defense Law Firm creado exitosamente"
    
    def get_context_function(self, request, form):
        return defense_law_firm_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateDefenseLawFirm(BaseUpdateView):
    model = DefenseLawFirm
    form_class = DefenseLawFirmForm
    success_url = reverse_lazy('defense-law-firm')
    template_name = 'show/show-defense-law-firm.html'
    success_message = "Defense Law Firm modificado exitosamente"
    
    def get_context_function(self, request, form):
        return defense_law_firm_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteDefenseLawFirm(BaseDeleteView):
    model = DefenseLawFirm
    redirect_url = 'defense-law-firm'
    template_name = 'show/show-defense-law-firm.html'
    success_message = "Defense Law Firm eliminado exitosamente"

    def get_context_function(self, request):
        return defense_law_firm_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreDefenseLawFirm(BaseRestoreView):
    model = DefenseLawFirm
    redirect_url = 'defense-law-firm'
    success_message = "Defense Law Firm restaurado exitosamente"


# ============================================================
# CLIENT
# ============================================================

def client_context(request, form=ClientForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Client',
        'form': form,
        'list_url': 'client',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = Client.objects.filter(
            Q(name__icontains=search) | 
            Q(id__icontains=search) |
            Q(address__icontains=search) |
            Q(ssn__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListClient(BaseListView):
    model = Client
    template_name = 'show/show-client.html'
    
    def get_context_function(self, request):
        return client_context(request)


@method_decorator(login_required, name='dispatch')
class CreateClient(BaseCreateView):
    model = Client
    form_class = ClientForm
    success_url = reverse_lazy('client')
    template_name = 'show/show-client.html'
    success_message = "Client creado exitosamente"
    
    def get_context_function(self, request, form):
        return client_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateClient(BaseUpdateView):
    model = Client
    form_class = ClientForm
    success_url = reverse_lazy('client')
    template_name = 'show/show-client.html'
    success_message = "Client modificado exitosamente"
    
    def get_context_function(self, request, form):
        return client_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteClient(BaseDeleteView):
    model = Client
    redirect_url = 'client'
    template_name = 'show/show-client.html'
    success_message = "Client eliminado exitosamente"

    def get_context_function(self, request):
        return client_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreClient(BaseRestoreView):
    model = Client
    redirect_url = 'client'
    success_message = "Client restaurado exitosamente"


# ============================================================
# CLAIM ADJUSTER
# ============================================================

def claim_adjuster_context(request, form=ClaimAdjusterForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Claim Adjuster',
        'form': form,
        'list_url': 'claim-adjuster',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = ClaimAdjuster.objects.filter(
            Q(name__icontains=search) | Q(id__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListClaimAdjuster(BaseListView):
    model = ClaimAdjuster
    template_name = 'show/show-claim-adjuster.html'
    
    def get_context_function(self, request):
        return claim_adjuster_context(request)


@method_decorator(login_required, name='dispatch')
class CreateClaimAdjuster(BaseCreateView):
    model = ClaimAdjuster
    form_class = ClaimAdjusterForm
    success_url = reverse_lazy('claim-adjuster')
    template_name = 'show/show-claim-adjuster.html'
    success_message = "Claim Adjuster creado exitosamente"
    
    def get_context_function(self, request, form):
        return claim_adjuster_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateClaimAdjuster(BaseUpdateView):
    model = ClaimAdjuster
    form_class = ClaimAdjusterForm
    success_url = reverse_lazy('claim-adjuster')
    template_name = 'show/show-claim-adjuster.html'
    success_message = "Claim Adjuster modificado exitosamente"
    
    def get_context_function(self, request, form):
        return claim_adjuster_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteClaimAdjuster(BaseDeleteView):
    model = ClaimAdjuster
    redirect_url = 'claim-adjuster'
    template_name = 'show/show-claim-adjuster.html'
    success_message = "Claim Adjuster eliminado exitosamente"

    def get_context_function(self, request):
        return claim_adjuster_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreClaimAdjuster(BaseRestoreView):
    model = ClaimAdjuster
    redirect_url = 'claim-adjuster'
    success_message = "Claim Adjuster restaurado exitosamente"


# ============================================================
# DEFENSE ATTORNEY
# ============================================================

def defense_attorney_context(request, form=DefenseAttorneyForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Defense Attorney',
        'form': form,
        'list_url': 'defense-attorney',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = DefenseAttorney.objects.filter(
            Q(name__icontains=search) | Q(id__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListDefenseAttorney(BaseListView):
    model = DefenseAttorney
    template_name = 'show/show-defense-attorney.html'
    
    def get_context_function(self, request):
        return defense_attorney_context(request)


@method_decorator(login_required, name='dispatch')
class CreateDefenseAttorney(BaseCreateView):
    model = DefenseAttorney
    form_class = DefenseAttorneyForm
    success_url = reverse_lazy('defense-attorney')
    template_name = 'show/show-defense-attorney.html'
    success_message = "Defense Attorney creado exitosamente"
    
    def get_context_function(self, request, form):
        return defense_attorney_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateDefenseAttorney(BaseUpdateView):
    model = DefenseAttorney
    form_class = DefenseAttorneyForm
    success_url = reverse_lazy('defense-attorney')
    template_name = 'show/show-defense-attorney.html'
    success_message = "Defense Attorney modificado exitosamente"
    
    def get_context_function(self, request, form):
        return defense_attorney_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteDefenseAttorney(BaseDeleteView):
    model = DefenseAttorney
    redirect_url = 'defense-attorney'
    template_name = 'show/show-defense-attorney.html'
    success_message = "Defense Attorney eliminado exitosamente"

    def get_context_function(self, request):
        return defense_attorney_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreDefenseAttorney(BaseRestoreView):
    model = DefenseAttorney
    redirect_url = 'defense-attorney'
    success_message = "Defense Attorney restaurado exitosamente"


# ============================================================
# DEFENSE ASSISTANT
# ============================================================

def defense_assistant_context(request, form=DefenseAssistantForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Defense Assistant',
        'form': form,
        'list_url': 'defense-assistant',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = DefenseAssistant.objects.filter(
            Q(name__icontains=search) | Q(id__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListDefenseAssistant(BaseListView):
    model = DefenseAssistant
    template_name = 'show/show-defense-assistant.html'
    
    def get_context_function(self, request):
        return defense_assistant_context(request)


@method_decorator(login_required, name='dispatch')
class CreateDefenseAssistant(BaseCreateView):
    model = DefenseAssistant
    form_class = DefenseAssistantForm
    success_url = reverse_lazy('defense-assistant')
    template_name = 'show/show-defense-assistant.html'
    success_message = "Defense Assistant creado exitosamente"
    
    def get_context_function(self, request, form):
        return defense_assistant_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateDefenseAssistant(BaseUpdateView):
    model = DefenseAssistant
    form_class = DefenseAssistantForm
    success_url = reverse_lazy('defense-assistant')
    template_name = 'show/show-defense-assistant.html'
    success_message = "Defense Assistant modificado exitosamente"
    
    def get_context_function(self, request, form):
        return defense_assistant_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteDefenseAssistant(BaseDeleteView):
    model = DefenseAssistant
    redirect_url = 'defense-assistant'
    template_name = 'show/show-defense-assistant.html'
    success_message = "Defense Assistant eliminado exitosamente"

    def get_context_function(self, request):
        return defense_assistant_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreDefenseAssistant(BaseRestoreView):
    model = DefenseAssistant
    redirect_url = 'defense-assistant'
    success_message = "Defense Assistant restaurado exitosamente"


# ============================================================
# INJURY TYPE
# ============================================================

def injury_type_context(request, form=InjuryTypeForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Injury Type',
        'form': form,
        'list_url': 'injury-type',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = InjuryType.objects.filter(
            Q(name__icontains=search) | Q(id__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context


@method_decorator(login_required, name='dispatch')
class ListInjuryType(BaseListView):
    model = InjuryType
    template_name = 'show/show-injury-type.html'
    
    def get_context_function(self, request):
        return injury_type_context(request)


@method_decorator(login_required, name='dispatch')
class CreateInjuryType(BaseCreateView):
    model = InjuryType
    form_class = InjuryTypeForm
    success_url = reverse_lazy('injury-type')
    template_name = 'show/show-injury-type.html'
    success_message = "Injury Type creado exitosamente"
    
    def get_context_function(self, request, form):
        return injury_type_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateInjuryType(BaseUpdateView):
    model = InjuryType
    form_class = InjuryTypeForm
    success_url = reverse_lazy('injury-type')
    template_name = 'show/show-injury-type.html'
    success_message = "Injury Type modificado exitosamente"
    
    def get_context_function(self, request, form):
        return injury_type_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteInjuryType(BaseDeleteView):
    model = InjuryType
    redirect_url = 'injury-type'
    template_name = 'show/show-injury-type.html'
    success_message = "Injury Type eliminado exitosamente"

    def get_context_function(self, request):
        return injury_type_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreInjuryType(BaseRestoreView):
    model = InjuryType
    redirect_url = 'injury-type'
    success_message = "Injury Type restaurado exitosamente"


# ============================================================
# BODY PART
# ============================================================

def body_part_context(request, form=BodyPartForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Body Part',
        'form': form,
        'list_url': 'body-part',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = BodyPart.objects.filter(
            Q(name__icontains=search) | Q(id__icontains=search)
        )
        
        # Filtro especial para BodyPart
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset

    return context


@method_decorator(login_required, name='dispatch')
class ListBodyPart(BaseListView):
    model = BodyPart
    template_name = 'show/show-body-part.html'
    
    def get_context_function(self, request):
        return body_part_context(request)


@method_decorator(login_required, name='dispatch')
class CreateBodyPart(BaseCreateView):
    model = BodyPart
    form_class = BodyPartForm
    success_url = reverse_lazy('body-part')
    template_name = 'show/show-body-part.html'
    success_message = "Body Part creado exitosamente"
    
    def get_context_function(self, request, form):
        return body_part_context(request, form)


@method_decorator(login_required, name='dispatch')
class UpdateBodyPart(BaseUpdateView):
    model = BodyPart
    form_class = BodyPartForm
    success_url = reverse_lazy('body-part')
    template_name = 'show/show-body-part.html'
    success_message = "Body Part modificado exitosamente"
    
    def get_context_function(self, request, form):
        return body_part_context(request, form)


@method_decorator(login_required, name='dispatch')
class DeleteBodyPart(BaseDeleteView):
    model = BodyPart
    redirect_url = 'body-part'
    template_name = 'show/show-body-part.html'
    success_message = "Body Part eliminado exitosamente"

    def get_context_function(self, request):
        return body_part_context(request)


@method_decorator(login_required, name='dispatch')
class RestoreBodyPart(BaseRestoreView):
    model = BodyPart
    redirect_url = 'body-part'
    success_message = "Body Part restaurado exitosamente"

# ============================================================
# INJURY (especial - tiene relaciones)
# ============================================================

def injury_context(request, form=InjuryForm(), return_query=True):
    search = request.GET.get('search', '')
    
    context = {
        'name': 'Injury',
        'form': form,
        'list_url': 'injury',
        'search': search,
        'create_active': False,
        'update_active': False,
        'delete_active': False,
        'restore_active': False,
    }
    
    if return_query:
        queryset = Injury.objects.filter(
            Q(id__icontains=search) |
            Q(date__icontains=search) |
            Q(part__name__icontains=search) |
            Q(type__name__icontains=search) |
            Q(case__id__icontains=search)
        )
        if request.user.systemuser.role != "admin":
            queryset = queryset.filter(active=True)
        context['items'] = queryset
    
    
    return context

@method_decorator(login_required, name='dispatch')
class ListInjury(BaseListView):
    model = Injury
    template_name = 'show/show-injury.html'
    
    def get_context_function(self, request):
        context = injury_context(request)
        context['body_parts'] = BodyPart.objects.filter(active=True)
        context['injury_types'] = InjuryType.objects.filter(active=True)
        return context


@method_decorator(login_required, name='dispatch')
class CreateInjury(BaseCreateView):
    model = Injury
    form_class = InjuryForm
    success_url = reverse_lazy('injury')
    template_name = 'show/show-injury.html'
    success_message = "Injury creado exitosamente"
    
    def get_context_function(self, request, form):
        context = injury_context(request, form)
        context['body_parts'] = BodyPart.objects.filter(active=True)
        context['injury_types'] = InjuryType.objects.filter(active=True)
        return context


@method_decorator(login_required, name='dispatch')
class UpdateInjury(BaseUpdateView):
    model = Injury
    form_class = InjuryForm
    success_url = reverse_lazy('injury')
    template_name = 'show/show-injury.html'
    success_message = "Injury modificado exitosamente"
    
    def get_context_function(self, request, form):
        context = injury_context(request, form)
        context['body_parts'] = BodyPart.objects.filter(active=True)
        context['injury_types'] = InjuryType.objects.filter(active=True)
        return context


@method_decorator(login_required, name='dispatch')
class DeleteInjury(BaseDeleteView):
    model = Injury
    redirect_url = 'injury'
    template_name = 'show/show-injury.html'
    success_message = "Injury eliminado exitosamente"

    def get_context_function(self, request, form):
        return injury_context(self, request)


@method_decorator(login_required, name='dispatch')
class RestoreInjury(BaseRestoreView):
    model = Injury
    redirect_url = 'injury'
    success_message = "Injury restaurado exitosamente"
 

from django.utils import timezone
from datetime import datetime, timedelta 
import pandas as pd

@method_decorator(login_required, name='dispatch')
class Report(View):

    def get(self, request):
        data = request.GET
        
        # Obtener fechas del request, con valores por defecto
        start_date_str = data.get('start_date', '')
        end_date_str = data.get('end_date', '')
        
        # Valores por defecto: una semana atrás y hoy
        today = timezone.now().date()
        default_start = today - timedelta(days=7)
        
        # Procesar start_date
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                start_date = default_start
        else:
            start_date = default_start
        
        # Procesar end_date
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                end_date = today
        else:
            end_date = today
         
        # Aplicar los mismos filtros que usas para mostrar
        cases = Case.objects.filter(
            updated_at__date__gte=start_date,
            updated_at__date__lte=end_date
        ).select_related('client', 'attorney', 'assistant', 'status')
        
        # Verificar si hay que exportar
        if data.get('export') == '1':  # El value="1" se captura aquí
            return self.export_to_excel(cases, start_date, end_date)
        
        # Si no, mostrar la vista normal
        cases = cases.order_by('-updated_at')
        
        context = {
            'items': cases,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
        }
        
        return render(request, 'report.html', context)
    
    def export_to_excel(self, cases, start_date, end_date):
        # Exporta los casos a un archivo Excel
        
        data = []
        for case in cases:
            # Convertir fechas timezone-aware a naive
            updated_at_naive = case.updated_at.replace(tzinfo=None) if case.updated_at else None
            created_at_naive = case.created_at.replace(tzinfo=None) if case.created_at else None
            settlement_date_naive = case.settlement_date  # DateField ya es naive
            adj_date_naive = case.adj_date  # DateField ya es naive
            date_assigned_naive = case.date_assigned if hasattr(case, 'date_assigned') else None  # DateField
            
            data.append({
                'Client Name': case.client.name,
                'Date Assigned': date_assigned_naive,
                'Attorney': case.attorney.name if case.attorney else '',
                'Assistant': case.assistant.name if case.assistant else '',
                'Case Status': case.status.name if case.status else '',
                'Details Registered': 'Yes' if case.def_lawfirm or case.claim_admin or case.def_attorney else 'No',
                'ADJ Filed': 'Yes' if case.adj_date else 'No',
                'ADJ Number': case.adj_num or '',
                'Claim Number': case.claim_num or '',
                'Claim Status': case.get_claim_status_display() or '',
                'Settled': 'Yes' if case.settlement_date else 'No',
                'Settlement Date': settlement_date_naive,
                'Active': 'Yes' if case.active else 'No',
                'Source': getattr(case, 'source', ''), 
            })
        
        df = pd.DataFrame(data)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'Cases_Report_Week-.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Cases Report', startrow=1)
            
            # Obtener el workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets['Cases Report']
            
            # Estilo para el encabezado
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplicar estilo a los encabezados (fila 2 porque startrow=1)
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Autoajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar la fila del encabezado
            worksheet.freeze_panes = 'A3'
        
        return response
    

class CaseHistory(View):

    def get(self, request, pk):  
        case = get_object_or_404(Case, id=pk)
        history = case.history.all().order_by('-history_date')
        
        # Obtener los verbose_name de los campos del modelo Case
        field_verbose_names = {}
        for field in Case._meta.get_fields():
            if hasattr(field, 'verbose_name'):
                field_verbose_names[field.name] = str(field.verbose_name)
            else:
                # Para campos que no tienen verbose_name, formatear el nombre
                field_verbose_names[field.name] = field.name.replace('_', ' ').title()
        
        # Diccionario adicional para campos personalizados (si necesitas sobrescribir)
        CUSTOM_NAMES = {
            'adj_num': 'Adjudication Number',
            'claim_num': 'Claim Number',
            'date_assigned': 'Date Assigned',
            'def_lawfirm': 'Defense Law Firm',
            'def_attorney': 'Defense Attorney',
            'def_assistant': 'Defense Assistant',
            'claim_admin': 'Claim Administrator',
            'claim_adjuster': 'Claim Adjuster',
        }
        
        # Combinar verbose_name con nombres personalizados (los personalizados tienen prioridad)
        for field_name, custom_name in CUSTOM_NAMES.items():
            field_verbose_names[field_name] = custom_name
        
        history_data = []
        history_list = list(history)
        
        for i, record in enumerate(history_list):
            changes = []
            
            # Comparar con el registro anterior si existe
            if i < len(history_list) - 1:
                prev_record = history_list[i + 1]
                diff = record.diff_against(prev_record)
                
                # Convertir los cambios a un formato usable en el template
                for change in diff.changes:
                    # ✅ Usar verbose_name en lugar del nombre del campo
                    field_display_name = field_verbose_names.get(change.field, change.field.replace('_', ' ').title())
                    
                    changes.append({
                        'field': field_display_name,
                        'field_name': change.field,  # Guardar también el nombre original por si acaso
                        'old': str(change.old) if change.old else '(none)',
                        'new': str(change.new) if change.new else '(none)',
                    })
            
            # Determinar el tipo de registro
            is_initial = (i == len(history_list) - 1)  # El último registro es el inicial
            has_changes = len(changes) > 0
            
            history_data.append({
                'date': record.history_date,
                'user': record.history_user.username if record.history_user else 'System',
                'changes': changes,
                'is_initial': is_initial,
                'has_changes': has_changes,
                'record_type': 'initial' if is_initial else ('update' if has_changes else 'empty_save'),
            })
        
        context = {
            'case': case,
            'history_data': history_data,
            'total_changes': len(history_data),
        }
        
        return render(request, 'case-history.html', context)

